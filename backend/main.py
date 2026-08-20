"""
Comparativa Quechua — Backend
FastAPI + PostgreSQL (asyncpg)
Los HTML se sirven dinámicamente desde endpoints para inyectar la URL de la API.
"""

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional
from contextlib import asynccontextmanager
from pathlib import Path
from bs4 import BeautifulSoup
import httpx, re, asyncpg, os, io
import openpyxl
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# ──────────────────────────────────────────────
# CONFIG — todo desde .env
# ──────────────────────────────────────────────
MOODLE_URL   = os.getenv("MOODLE_URL",   "https://tu-moodle.edu.pe")
MOODLE_TOKEN = os.getenv("MOODLE_TOKEN", "TU_TOKEN_WEBSERVICE")
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://user:password@localhost:5432/quechua_db")
API_HOST     = os.getenv("API_HOST",     "http://localhost")
API_PORT     = os.getenv("API_PORT",     "9005")
API_BASE     = f"{API_HOST}:{API_PORT}"


# ──────────────────────────────────────────────
# BASE DE DATOS
# ──────────────────────────────────────────────
pool: asyncpg.Pool = None

async def init_db():
    await pool.execute("""
        CREATE TABLE IF NOT EXISTS qch_actividades (
            quizid           INTEGER PRIMARY KEY,
            cmid             INTEGER  NOT NULL,
            curid            INTEGER  NOT NULL,
            numero_eval      SMALLINT NOT NULL CHECK (numero_eval BETWEEN 1 AND 8),
            numero_intento   SMALLINT NOT NULL CHECK (numero_intento IN (1, 2)),
            nombre           TEXT,
            texto_correcto   TEXT     NOT NULL,
            creado_en        TIMESTAMPTZ DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS qch_resultados (
            id                 SERIAL PRIMARY KEY,
            quizid             INTEGER      NOT NULL,
            cmid               INTEGER      NOT NULL,
            curid              INTEGER      NOT NULL,
            id_user            INTEGER      NOT NULL,
            numero_eval        SMALLINT     NOT NULL,
            numero_intento     SMALLINT     NOT NULL,
            puntaje            SMALLINT     NOT NULL CHECK (puntaje IN (0, 1)),
            aciertos           INTEGER      NOT NULL DEFAULT 0,
            porcentaje_acierto NUMERIC(5,2) NOT NULL DEFAULT 0,
            texto_docente      TEXT,
            fecha              TIMESTAMPTZ  DEFAULT NOW(),
            CONSTRAINT uq_resultado UNIQUE (quizid, id_user)
        );

        CREATE INDEX IF NOT EXISTS idx_resultados_user_curid
            ON qch_resultados (id_user, curid);
    """)

@asynccontextmanager
async def lifespan(app: FastAPI):
    global pool
    pool = await asyncpg.create_pool(DATABASE_URL, min_size=2, max_size=10)
    await init_db()
    yield
    await pool.close()


# ──────────────────────────────────────────────
# APP
# ──────────────────────────────────────────────
app = FastAPI(title="Comparativa Quechua", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=os.getenv("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


# ──────────────────────────────────────────────
# MODELOS
# ──────────────────────────────────────────────
class Actividad(BaseModel):
    quizid:         int
    cmid:           int
    curid:          int
    numero_eval:    int
    numero_intento: int
    nombre:         Optional[str] = None
    texto_correcto: str

class ComparativaRequest(BaseModel):
    quizid:  int
    cmid:    int
    curid:   int
    id_user: int


# ──────────────────────────────────────────────
# LÓGICA DE COMPARACIÓN
# ──────────────────────────────────────────────
def normalizar(texto: str) -> list[str]:
    texto = texto.lower()
    # Eliminar puntuación EXCEPTO apóstrofe (') y saltillo (`)
    # que son vitales en lenguas originarias (quechua, aymara)
    texto = re.sub(r"[.,;:¡!¿?\"()\[\]{}«»\-]", "", texto)
    texto = re.sub(r"[\n\r\t]", " ", texto)
    texto = re.sub(r"\s+", " ", texto)
    return [w for w in texto.strip().split() if w]

def comparar(texto_correcto: str, texto_docente: str) -> dict:
    from difflib import SequenceMatcher
    esperado = normalizar(texto_correcto)
    recibido = normalizar(texto_docente)
    total    = len(esperado)

    matcher  = SequenceMatcher(None, esperado, recibido)
    aciertos = sum(i2 - i1 for tag, i1, i2, j1, j2 in matcher.get_opcodes() if tag == "equal")

    porcentaje = round((aciertos / total * 100), 2) if total > 0 else 0.0
    puntaje    = 1 if aciertos == total else 0

    return {
        "puntaje":            puntaje,
        "aciertos":           aciertos,
        "porcentaje_acierto": porcentaje,
    }


# ──────────────────────────────────────────────
# MOODLE API
# ──────────────────────────────────────────────
def _moodle_params(wsfunction: str, extra: dict) -> dict:
    return {"wstoken": MOODLE_TOKEN, "wsfunction": wsfunction,
            "moodlewsrestformat": "json", **extra}

async def _moodle_get(client: httpx.AsyncClient, wsfunction: str, extra: dict) -> dict:
    r = await client.get(
        f"{MOODLE_URL}/webservice/rest/server.php",
        params=_moodle_params(wsfunction, extra)
    )
    r.raise_for_status()
    data = r.json()
    if "exception" in data:
        raise HTTPException(502, f"Moodle [{wsfunction}]: {data.get('message')}")
    return data

async def obtener_texto_moodle(quizid: int, id_user: int) -> str:
    async with httpx.AsyncClient(timeout=30, verify=False) as client:
        attempts_data = await _moodle_get(client, "mod_quiz_get_user_attempts",
                                          {"quizid": quizid, "userid": id_user, "status": "all"})
        attempts = attempts_data.get("attempts", [])
        if not attempts:
            raise HTTPException(404,
                f"No se encontraron intentos para userid={id_user} en quizid={quizid}")

        finished   = [a for a in attempts if a.get("state") == "finished"]
        inprogress = [a for a in attempts if a.get("state") == "inprogress"]

        if finished:
            attempt = sorted(finished, key=lambda x: x.get("timemodified", 0))[-1]
        elif inprogress:
            attempt = sorted(inprogress, key=lambda x: x.get("timemodified", 0))[-1]
        else:
            raise HTTPException(422,
                f"El usuario userid={id_user} no tiene intentos finalizados en quizid={quizid}")

        review = await _moodle_get(client, "mod_quiz_get_attempt_review",
                                   {"attemptid": attempt["id"], "page": -1})

    for question in review.get("questions", []):
        if question.get("type") == "essay":
            html = question.get("html", "")
            if html:
                try:
                    soup = BeautifulSoup(html, "html.parser")
                    candidatos = []

                    # Fuente 1: textarea con clase específica de essay
                    textarea = soup.find("textarea", class_="qtype_essay_response")
                    if not textarea:
                        textarea = soup.find("textarea", attrs={"readonly": True})
                    if textarea:
                        t = (textarea.string or textarea.get_text(strip=True)).strip()
                        t = re.sub(r"^Texto de la respuesta Pregunta\s*\d+\s*", "", t).strip()
                        if t:
                            candidatos.append(t)

                    # Fuente 2: historial "Guardada:" — split robusto para multilínea
                    texto_plano = soup.get_text(" ", strip=True)
                    if "Guardada:" in texto_plano:
                        t = texto_plano.split("Guardada:")[-1].strip()
                        for corte in ["Respuesta guardada", "Intento finalizado"]:
                            if corte in t:
                                t = t.split(corte)[0].strip()
                        if t:
                            candidatos.append(t)

                    # Fuente 3: div de respuesta essay
                    answer_div = soup.find("div", class_=["answer", "qtype_essay_response"])
                    if answer_div:
                        t = answer_div.get_text(" ", strip=True)
                        if t:
                            candidatos.append(t)

                    # Limpiar prefijo de accesibilidad y quedarse con el más largo
                    if candidatos:
                        candidatos = [
                            re.sub(r"^Texto de la respuesta Pregunta\s*\d+\s*", "", c).strip()
                            for c in candidatos
                        ]
                        candidatos = [c for c in candidatos if c]
                        if candidatos:
                            return max(candidatos, key=len)

                except Exception as e:
                    print(f"[ERROR] BeautifulSoup: {e}")

            # Fallback: responsesummary
            texto = (question.get("responsesummary") or "").strip()
            if texto:
                return texto

    return ""


# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def row_to_dict(row) -> dict:
    return dict(row) if row else None

def build_respuesta(act: dict, res: dict) -> dict:
    es_intento2 = act["numero_intento"] == 2
    return {
        "existe":         res is not None,
        "puntaje":        res["puntaje"]       if res else None,
        "texto_docente":  res["texto_docente"] if res else None,
        "fecha":          res["fecha"].isoformat() if res and res["fecha"] else None,
        "texto_correcto": act["texto_correcto"] if (res and es_intento2) else None,
        "numero_eval":    act["numero_eval"],
        "numero_intento": act["numero_intento"],
        "nombre":         act["nombre"],
        "curid":          act["curid"],
    }


# ──────────────────────────────────────────────
# ENDPOINTS — PÁGINAS HTML
# ──────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def pagina_resultado():
    html = Path(__file__).parent.parent / "frontend" / "index.html"
    content = html.read_text(encoding="utf-8")
    content = content.replace("__API_BASE__", API_BASE)
    return HTMLResponse(content)

@app.get("/gestor", response_class=HTMLResponse)
async def pagina_gestor():
    html = Path(__file__).parent.parent / "frontend" / "gestor.html"
    content = html.read_text(encoding="utf-8")
    content = content.replace("__API_BASE__", API_BASE)
    return HTMLResponse(content)


# ──────────────────────────────────────────────
# ENDPOINTS — GESTOR (qch_actividades)
# ──────────────────────────────────────────────
@app.get("/api/actividades")
async def listar_actividades(curid: Optional[int] = None):
    async with pool.acquire() as conn:
        if curid:
            rows = await conn.fetch(
                "SELECT * FROM qch_actividades WHERE curid=$1 ORDER BY numero_eval, numero_intento",
                curid
            )
        else:
            rows = await conn.fetch(
                "SELECT * FROM qch_actividades ORDER BY curid, numero_eval, numero_intento"
            )
    return [row_to_dict(r) for r in rows]

@app.post("/api/actividades", status_code=201)
async def crear_actividad(act: Actividad):
    async with pool.acquire() as conn:
        try:
            await conn.execute(
                """INSERT INTO qch_actividades
                       (quizid, cmid, curid, numero_eval, numero_intento, nombre, texto_correcto)
                   VALUES ($1,$2,$3,$4,$5,$6,$7)""",
                act.quizid, act.cmid, act.curid,
                act.numero_eval, act.numero_intento,
                act.nombre, act.texto_correcto
            )
        except asyncpg.UniqueViolationError:
            raise HTTPException(409, f"Ya existe una actividad con quizid={act.quizid}")
        except asyncpg.CheckViolationError as e:
            raise HTTPException(422, f"Valor fuera de rango: {e}")
    return {"ok": True}

@app.put("/api/actividades/{quizid}")
async def actualizar_actividad(quizid: int, act: Actividad):
    async with pool.acquire() as conn:
        result = await conn.execute(
            """UPDATE qch_actividades
               SET cmid=$1, curid=$2, numero_eval=$3, numero_intento=$4,
                   nombre=$5, texto_correcto=$6
               WHERE quizid=$7""",
            act.cmid, act.curid, act.numero_eval, act.numero_intento,
            act.nombre, act.texto_correcto, quizid
        )
    if result == "UPDATE 0":
        raise HTTPException(404, f"quizid={quizid} no encontrado")
    return {"ok": True}

@app.delete("/api/actividades/{quizid}")
async def eliminar_actividad(quizid: int):
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM qch_actividades WHERE quizid=$1", quizid)
    return {"ok": True}


# ──────────────────────────────────────────────
# ENDPOINT — DESCARGAR PLANTILLA EXCEL
# ──────────────────────────────────────────────
@app.get("/api/actividades/plantilla")
async def descargar_plantilla():
    """Descarga un Excel vacío con las columnas correctas como plantilla."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actividades"

    headers = ["quizid", "cmid", "curid", "numero_eval", "numero_intento", "nombre", "texto_correcto"]
    ws.append(headers)

    # Fila de ejemplo
    ws.append([12192, 206332, 2679, 1, 1, "U2S1 P1 – Primer intento", "Texto correcto aquí..."])

    # Ajustar ancho de columnas
    anchos = [10, 10, 10, 13, 15, 30, 50]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_actividades.xlsx"}
    )


# ──────────────────────────────────────────────
# ENDPOINT — CARGA MASIVA DESDE EXCEL
# ──────────────────────────────────────────────
@app.post("/api/actividades/carga-masiva")
async def carga_masiva(file: UploadFile = File(...)):
    """
    Sube un Excel (.xlsx) con columnas:
    quizid | cmid | curid | numero_eval | numero_intento | nombre | texto_correcto
    Inserta o actualiza (upsert) cada fila.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx")

    contenido = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb.active

    # Leer encabezados de la primera fila
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    cols_requeridas = {"quizid", "cmid", "curid", "numero_eval", "numero_intento", "texto_correcto"}
    cols_faltantes  = cols_requeridas - set(headers)
    if cols_faltantes:
        raise HTTPException(400, f"Columnas faltantes en el Excel: {', '.join(cols_faltantes)}")

    idx = {h: i for i, h in enumerate(headers)}

    insertados   = 0
    actualizados = 0
    errores      = []

    async with pool.acquire() as conn:
        for num_fila, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue
            try:
                quizid         = int(row[idx["quizid"]])
                cmid           = int(row[idx["cmid"]])
                curid          = int(row[idx["curid"]])
                numero_eval    = int(row[idx["numero_eval"]])
                numero_intento = int(row[idx["numero_intento"]])
                nombre = str(row[idx["nombre"]]).strip().replace("_x000D_", "").strip() if "nombre" in idx and row[idx["nombre"]] else None
                texto_correcto = str(row[idx["texto_correcto"]]).strip()
                texto_correcto = texto_correcto.replace("_x000D_", "\n").strip()

                if not texto_correcto:
                    errores.append(f"Fila {num_fila}: texto_correcto vacío")
                    continue

                existente = await conn.fetchrow(
                    "SELECT quizid FROM qch_actividades WHERE quizid=$1", quizid
                )
                if existente:
                    await conn.execute(
                        """UPDATE qch_actividades
                           SET cmid=$1, curid=$2, numero_eval=$3, numero_intento=$4,
                               nombre=$5, texto_correcto=$6
                           WHERE quizid=$7""",
                        cmid, curid, numero_eval, numero_intento,
                        nombre, texto_correcto, quizid
                    )
                    actualizados += 1
                else:
                    await conn.execute(
                        """INSERT INTO qch_actividades
                               (quizid, cmid, curid, numero_eval, numero_intento, nombre, texto_correcto)
                           VALUES ($1,$2,$3,$4,$5,$6,$7)""",
                        quizid, cmid, curid, numero_eval, numero_intento,
                        nombre, texto_correcto
                    )
                    insertados += 1

            except Exception as e:
                errores.append(f"Fila {num_fila}: {str(e)}")

    return {
        "ok":          True,
        "insertados":  insertados,
        "actualizados": actualizados,
        "errores":     errores,
    }


# ──────────────────────────────────────────────
# ENDPOINTS — RESULTADO (qch_resultados)
# ──────────────────────────────────────────────
@app.get("/api/resultado/{quizid}/{id_user}")
async def obtener_resultado(quizid: int, id_user: int):
    async with pool.acquire() as conn:
        act = row_to_dict(await conn.fetchrow(
            "SELECT * FROM qch_actividades WHERE quizid=$1", quizid
        ))
        if not act:
            raise HTTPException(404, f"quizid={quizid} no configurado en el gestor")
        res = row_to_dict(await conn.fetchrow(
            "SELECT * FROM qch_resultados WHERE quizid=$1 AND id_user=$2", quizid, id_user
        ))
    return build_respuesta(act, res)

@app.post("/api/comparar")
async def comparar_y_guardar(req: ComparativaRequest):
    async with pool.acquire() as conn:
        act = row_to_dict(await conn.fetchrow(
            "SELECT * FROM qch_actividades WHERE quizid=$1", req.quizid
        ))
        if not act:
            raise HTTPException(404, f"quizid={req.quizid} no configurado en el gestor")

        existente = row_to_dict(await conn.fetchrow(
            "SELECT * FROM qch_resultados WHERE quizid=$1 AND id_user=$2",
            req.quizid, req.id_user
        ))
        if existente:
            return {"ya_existia": True, **build_respuesta(act, existente)}

    texto_docente = await obtener_texto_moodle(req.quizid, req.id_user)
    if not texto_docente.strip():
        texto_docente = ""
        cmp = {"puntaje": 0, "aciertos": 0, "porcentaje_acierto": 0.0}
    else:
        cmp = comparar(act["texto_correcto"], texto_docente)

    async with pool.acquire() as conn:
        await conn.execute(
            """INSERT INTO qch_resultados
                   (quizid, cmid, curid, id_user, numero_eval, numero_intento,
                    puntaje, aciertos, porcentaje_acierto, texto_docente)
               VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)""",
            req.quizid, req.cmid, req.curid,
            req.id_user, act["numero_eval"], act["numero_intento"],
            cmp["puntaje"], cmp["aciertos"], cmp["porcentaje_acierto"], texto_docente
        )
        res = row_to_dict(await conn.fetchrow(
            "SELECT * FROM qch_resultados WHERE quizid=$1 AND id_user=$2",
            req.quizid, req.id_user
        ))
    return {"ya_existia": False, **build_respuesta(act, res)}


# ──────────────────────────────────────────────
# ENDPOINT — RESUMEN ADMIN
# ──────────────────────────────────────────────
@app.get("/api/resumen/{curid}")
async def resumen_curso(curid: int):
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """SELECT r.id_user, r.numero_eval, r.numero_intento,
                      r.puntaje, r.aciertos, r.porcentaje_acierto,
                      r.texto_docente, r.fecha
               FROM qch_resultados r
               WHERE r.curid = $1
               ORDER BY r.id_user, r.numero_eval, r.numero_intento""",
            curid
        )
    return [
        {**row_to_dict(r), "fecha": r["fecha"].isoformat() if r["fecha"] else None}
        for r in rows
    ]
